"""Tests for revision features: fold-isolated normalization, SHAP-based candidate
selection, per-fold panel overlap, and the fold-overlap export sheet."""

from __future__ import annotations

import io
from types import SimpleNamespace

import numpy as np
import pytest

from mlheatmap.core.normalization import (
    FoldNormalizer,
    deseq2_normalize,
    log2_normalize,
    tpm_normalize,
)
from mlheatmap.core.biomarker import run_biomarker_analysis


def _toy_counts(seed=0, n_genes=120, per_group=10, n_groups=3):
    rng = np.random.default_rng(seed)
    blocks, labels = [], []
    for gi in range(n_groups):
        base = rng.gamma(2.0, 40.0, size=(n_genes, 1))
        signal = np.zeros((n_genes, 1))
        signal[gi * 8:(gi + 1) * 8] = 250.0
        mat = rng.poisson((base + signal) * np.ones((n_genes, per_group)) + 1).astype(float)
        blocks.append(mat)
        labels += [f"G{gi}"] * per_group
    counts = np.concatenate(blocks, axis=1)
    groups = {}
    for idx, lab in enumerate(labels):
        groups.setdefault(lab, []).append(idx)
    gene_names = [f"gene{i}" for i in range(n_genes)]
    return counts, gene_names, groups


class TestFoldNormalizer:
    def test_global_matches_fit_transform(self):
        counts, _, _ = _toy_counts()
        glob = deseq2_normalize(counts)
        fold = FoldNormalizer("deseq2").fit_transform(counts)
        assert np.allclose(glob, fold, equal_nan=True)

    @pytest.mark.parametrize("method", ["deseq2", "tpm", "log2"])
    def test_fit_transform_shape_and_finiteness(self, method):
        counts, _, _ = _toy_counts()
        fn = FoldNormalizer(method).fit(counts[:, :20])
        out = fn.transform(counts[:, 20:])
        assert out.shape == counts[:, 20:].shape
        assert np.isfinite(out).all()

    def test_tpm_log2_fold_equal_global(self):
        counts, _, _ = _toy_counts()
        assert np.allclose(tpm_normalize(counts), FoldNormalizer("tpm").fit_transform(counts), equal_nan=True)
        assert np.allclose(log2_normalize(counts), FoldNormalizer("log2").fit_transform(counts), equal_nan=True)

    def test_unknown_method_raises(self):
        with pytest.raises(ValueError):
            FoldNormalizer("bogus")

    def test_transform_before_fit_raises(self):
        with pytest.raises(RuntimeError):
            FoldNormalizer("deseq2").transform(np.ones((5, 5)))


class TestBiomarkerOptions:
    def _run(self, **kw):
        counts, gene_names, groups = _toy_counts()
        normalized = deseq2_normalize(counts)
        return run_biomarker_analysis(
            expression=normalized, gene_names=gene_names, sample_groups=groups,
            n_top_genes=12, n_estimators=60, cv_folds=2, model="rf", panel_method="forward",
            raw_counts=counts, **kw,
        )

    def test_global_default_scope_and_fold_panels(self):
        res = self._run()
        assert res["normalization_scope"] == "global"
        assert res["selection_basis"] == "importance"
        assert res["optimal_combo"]["fold_panels"]
        assert all("genes" in fp and "fold" in fp for fp in res["optimal_combo"]["fold_panels"])

    def test_fold_isolated_normalization(self):
        res = self._run(per_fold_normalize=True, norm_method="deseq2")
        assert res["normalization_scope"] == "fold_isolated"
        assert res["norm_method"] == "deseq2"
        assert 0.0 <= res["accuracy"] <= 1.0

    def test_per_fold_requires_raw_counts(self):
        counts, gene_names, groups = _toy_counts()
        normalized = deseq2_normalize(counts)
        with pytest.raises(ValueError):
            run_biomarker_analysis(
                expression=normalized, gene_names=gene_names, sample_groups=groups,
                n_estimators=40, cv_folds=2, per_fold_normalize=True, raw_counts=None,
            )

    def test_shap_selection_basis(self):
        res = self._run(selection_basis="shap")
        assert res["selection_basis"] == "shap"

    def test_invalid_selection_basis(self):
        counts, gene_names, groups = _toy_counts()
        with pytest.raises(ValueError):
            run_biomarker_analysis(
                expression=deseq2_normalize(counts), gene_names=gene_names, sample_groups=groups,
                n_estimators=40, cv_folds=2, selection_basis="nonsense",
            )


class TestFoldOverlapExport:
    def test_export_includes_fold_overlap_sheet(self):
        import openpyxl

        from mlheatmap.core.export import export_results_excel

        counts, gene_names, groups = _toy_counts()
        normalized = deseq2_normalize(counts)
        res = run_biomarker_analysis(
            expression=normalized, gene_names=gene_names, sample_groups=groups,
            n_top_genes=12, n_estimators=60, cv_folds=2, model="rf", panel_method="forward",
        )
        sample_names = [f"s{i}" for i in range(counts.shape[1])]
        session = SimpleNamespace(
            id="test", species="human", id_type="symbol", norm_method="deseq2",
            deg_effect_size_basis="size_factor_normalized_counts",
            gene_names=gene_names, sample_names=sample_names, excluded_samples=[],
            groups={g: [sample_names[i] for i in idx] for g, idx in groups.items()},
            metadata={}, normalized=normalized, biomarker_results=res, deg_results=None,
        )
        data = export_results_excel(session, include_normalized_expression=False)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert "Panel Fold Overlap" in wb.sheetnames
        rows = list(wb["Panel Fold Overlap"].iter_rows(values_only=True))
        assert rows and rows[0][0] == "gene"
