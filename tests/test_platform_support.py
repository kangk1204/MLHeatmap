"""Cross-platform packaging and runtime support tests."""

from __future__ import annotations

import json
import subprocess
import sys
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from fastapi.testclient import TestClient


DATA_DIR = Path(__file__).parent / "data"


def _prepare_session(client: TestClient) -> str:
    upload_path = DATA_DIR / "small_quick_test.csv"
    with upload_path.open("rb") as handle:
        response = client.post("/api/v1/upload", files={"file": ("test.csv", handle, "text/csv")})
    assert response.status_code == 200
    session_id = response.json()["session_id"]
    samples = response.json()["sample_names"]

    response = client.post(
        "/api/v1/gene-mapping",
        json={"session_id": session_id, "species": "human", "id_type": "auto"},
    )
    assert response.status_code == 200

    response = client.post(
        "/api/v1/normalize",
        json={"session_id": session_id, "method": "log2"},
    )
    assert response.status_code == 200

    half = len(samples) // 2
    response = client.post(
        "/api/v1/groups",
        json={
            "session_id": session_id,
            "groups": {"Control": samples[:half], "Case": samples[half:]},
        },
    )
    assert response.status_code == 200
    return session_id


def _client() -> TestClient:
    from mlheatmap.server import create_app

    return TestClient(create_app())


def test_capabilities_endpoint_reports_core_runtime():
    client = _client()
    response = client.get("/api/v1/capabilities")
    assert response.status_code == 200

    data = response.json()
    assert data["python"]["supported"] == ">=3.11,<3.13"
    assert data["gene_tables"]["human"] is True
    assert data["gene_tables"]["mouse"] is True
    assert data["exports"]["image_mode"] == "browser"
    assert data["exports"]["results_excel"] is True
    assert data["models"]["rf"]["available"] is True


def test_app_limits_cors_to_local_origins():
    from mlheatmap.server import LOCAL_ORIGIN_REGEX, create_app

    app = create_app()
    cors = next(middleware for middleware in app.user_middleware if middleware.cls.__name__ == "CORSMiddleware")

    assert cors.kwargs["allow_origins"] == []
    assert cors.kwargs["allow_origin_regex"] == LOCAL_ORIGIN_REGEX


def test_packaged_gene_tables_exist():
    from importlib import resources

    from mlheatmap.core.gene_mapping import has_gene_table

    assert resources.files("mlheatmap.data").joinpath("human_genes.tsv").is_file()
    assert resources.files("mlheatmap.data").joinpath("mouse_genes.tsv").is_file()
    assert has_gene_table("human") is True
    assert has_gene_table("mouse") is True


def test_export_endpoint_rejects_server_side_image_exports():
    client = _client()
    session_id = _prepare_session(client)

    response = client.get(f"/api/v1/export?session_id={session_id}&type=heatmap_png")
    assert response.status_code == 400
    assert "browser" in response.json()["error"].lower()


def test_biomarker_stream_rejects_unavailable_optional_models():
    client = _client()
    session_id = _prepare_session(client)

    capability = {
        "id": "xgboost",
        "label": "XGBoost",
        "available": False,
        "known": True,
        "install_profile": "full",
        "unavailable_reason": "XGBoost is not available in this installation.",
    }
    with patch("mlheatmap.api.biomarker.get_model_capability", return_value=capability):
        response = client.get(f"/api/v1/biomarker/stream?session_id={session_id}&model=xgboost")

    assert response.status_code == 400
    assert "not available" in response.json()["error"].lower()


def test_heatmap_server_render_returns_png_bytes():
    client = _client()
    session_id = _prepare_session(client)

    response = client.get(f"/api/v1/heatmap/render?session_id={session_id}&top_n=20&fmt=png")
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("image/png")
    assert len(response.content) > 1000


def test_browser_export_script_uses_client_side_plot_export():
    export_js = (Path(__file__).parents[1] / "src" / "mlheatmap" / "static" / "js" / "export.js").read_text(
        encoding="utf-8"
    )

    assert "Plotly.downloadImage" in export_js
    assert "_exportPlotlyElement" in export_js
    assert "_downloadServerHeatmap" in export_js
    assert "results_excel" in export_js


def test_excel_export_includes_metadata_sheet():
    from openpyxl import load_workbook

    client = _client()
    session_id = _prepare_session(client)

    client.get(f"/api/v1/biomarker/deg?session_id={session_id}&reference_group=Control")
    client.get(f"/api/v1/biomarker/stream?session_id={session_id}&n_top_genes=5&n_estimators=50&cv_folds=2")
    response = client.get(f"/api/v1/export?session_id={session_id}&type=results_excel")

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), read_only=True)
    assert "Metadata" in workbook.sheetnames
    metadata_rows = list(workbook["Metadata"].iter_rows(values_only=True))
    assert any(row[0] == "app.version" for row in metadata_rows if row and row[0])
    assert any(row[0] == "metadata.normalization.effect_size_basis" for row in metadata_rows if row and row[0])


def test_paper_reproduce_script_writes_expected_outputs(tmp_path):
    groups_path = tmp_path / "groups.json"
    groups_path.write_text(
        json.dumps(
            {
                "Control": ["Ctrl_1", "Ctrl_2", "Ctrl_3"],
                "Case": ["Drug_1", "Drug_2", "Drug_3"],
            }
        ),
        encoding="utf-8",
    )
    output_dir = tmp_path / "out"

    proc = subprocess.run(
        [
            sys.executable,
            "scripts/paper_reproduce.py",
            "--input",
            str(DATA_DIR / "small_quick_test.csv"),
            "--groups-json",
            str(groups_path),
            "--output-dir",
            str(output_dir),
            "--species",
            "human",
            "--id-type",
            "auto",
            "--normalize",
            "log2",
            "--n-top-genes",
            "5",
            "--n-estimators",
            "50",
            "--cv-folds",
            "2",
        ],
        cwd=Path(__file__).parents[1],
        check=True,
        capture_output=True,
        text=True,
    )

    assert proc.returncode == 0
    assert (output_dir / "biomarker.json").is_file()
    assert (output_dir / "deg.json").is_file()
    assert (output_dir / "metadata.json").is_file()
    assert (output_dir / "results.xlsx").is_file()
    metadata = json.loads((output_dir / "metadata.json").read_text(encoding="utf-8"))
    assert metadata["metadata"]["reproducibility"]["git_commit"] is not None


def test_optional_module_runtime_errors_do_not_break_capability_checks():
    from mlheatmap.core import capabilities

    capabilities._module_available.cache_clear()
    try:
        with patch("mlheatmap.core.capabilities.importlib.import_module", side_effect=OSError("libgomp.so.1 missing")):
            available, reason = capabilities._module_available("lightgbm")

        assert available is False
        assert "libgomp" in reason
    finally:
        capabilities._module_available.cache_clear()
